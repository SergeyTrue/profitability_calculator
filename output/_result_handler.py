from pathlib import Path
from typing import Any, final, Iterable, Sequence
from abc import ABC, abstractmethod

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook

from ..common import SuccessUserResult, BadResult, UnknownUserResult, ConcatenatedUserResult, NoExcelFilesResult
from mc.common import User


class IResultHandler(ABC):
    _handled_result_cls = None

    def __init__(self):
        self._current_message = None

    @final
    def handle(self, message: Any):
        self._current_message = message
        if isinstance(message.data, self._handled_result_cls):
            res = self._handle(message.data)
            self._current_message = None
            return res

        raise RuntimeError(message)

    @abstractmethod
    def _handle(self, message: Any):
        pass

    def __init_subclass__(cls, **kwargs):
        assert cls._handled_result_cls
        assert isinstance(cls._handled_result_cls, type)


class UnknownUserResultHandler(IResultHandler):
    _handled_result_cls = UnknownUserResult

    def __init__(self, output_file: Path):
        super().__init__()
        self._output_file = output_file

    def _handle(self, result: UnknownUserResult):
        with self._output_file.open('w', encoding='utf-8') as stream:
            stream.write(result.content)


class BadResultHandler(IResultHandler):
    _handled_result_cls = BadResult

    def __init__(self, output_file: Path):
        super().__init__()
        self._output_file = output_file

    def _handle(self, result: BadResult):
        with self._output_file.open('a', encoding='utf-8') as stream:
            for file, data in result.wrong.items():
                stream.write(f'<p>Файл:  {file.stem},</p>')

                wrong_price = data.get('wrong_price')
                if wrong_price is not None and not wrong_price.empty:
                    stream.write('<p>Цена должна быть числом. У этих артикулов цена выражается не числом:</p>')
                    stream.write(wrong_price.to_html(index=False, justify='center'))

                empty_quantity = data.get('empty_qty')
                if empty_quantity is not None and not empty_quantity.empty:
                    stream.write('<p>Не указано количество у этих артикулов:</p>')
                    stream.write(empty_quantity.to_html(index=False, justify='center'))
                    stream.write('<p></p>')

                wrong_quantity = data.get('wrong_quantity')
                if wrong_quantity is not None and not wrong_quantity.empty:
                    stream.write('<p>Количество, указанное у этих артикулов, не является числом:</p>')
                    stream.write(wrong_quantity.to_html(index=False, justify='center'))

                bad_articles = data.get('bad_articles')
                if bad_articles is not None and not bad_articles.empty:
                    stream.write('<p>Неопознанные артикулы. Пожалуйста, укажите цену на них:</p>')
                    stream.write(bad_articles.to_html(index=False, justify='center'))


class SuccessAdminHeartBeatResultHandler(IResultHandler):
    _handled_result_cls = SuccessUserResult

    def __init__(self, bad_cost_file: Path,
                 bad_cost_html_file: Path,
                 full_emulator_files: Sequence[Iterable[Path]],
                 equal_margin_xls: Sequence[Iterable[Path]]):
        super().__init__()
        self._bad_cost_file = bad_cost_file
        self._bad_cost_html_file = bad_cost_html_file
        self._full_emulator_files = list(full_emulator_files)
        self._equal_margin_xls = list(equal_margin_xls)

    def _handle(self, result: SuccessUserResult):
        for child in result.children:
            if (df := child.bad_cost) is not None and not df.empty:
                df.to_csv(str(self._bad_cost_file), header=False, sep='\t', index=False, mode='a')
                with self._bad_cost_html_file.open('a', encoding='utf-8') as stream:
                    stream.write(df.to_html(index=False, justify='center'))

            if ratio := child.bad_cost_ratio:
                with self._bad_cost_html_file.open('a', encoding='utf-8') as stream:
                    stream.write(str(ratio))

            if (df := child.equal_margin_df) is not None and not df.empty:
                files = self._equal_margin_xls.pop(0)
                for file in files:
                    with pd.ExcelWriter(str(file), engine='openpyxl', mode='w') as writer:
                        df.to_excel(writer, sheet_name='Сводная', index=False, startrow=3)
                    _fill_cells_for_eq_margin(self._current_message.sender.company, file)

            if (df := child.full_emulator) is not None and not df.empty:
                files = self._full_emulator_files.pop(0)
                for file in files:
                    df.to_csv(str(file), header=False, sep='\t', index=False, mode='a')


class SuccessUserResultHandler(IResultHandler):
    _handled_result_cls = SuccessUserResult

    def __init__(self, reply_file: Path, equal_margin_xls: Sequence[Iterable[Path]]):
        super().__init__()
        self.reply_file = reply_file
        self._equal_margin_xls = list(equal_margin_xls)

    def _handle(self, result: SuccessUserResult):
        with self.reply_file.open('w', encoding='utf-8') as stream:

            stream.write(f'<header>{self._current_message.sender.name},</header>\n')
            stream.write('<p style = "color:red" > Конфиденциально.  </p>')
            if self._current_message.sender.company != 'RU41':
                stream.write(f'<p> Расчёт для компании {self._current_message.sender.company}</p>')

            for child in result.children:
                stream.write(f'<p>Файл:  {child.path.stem}</p>')

                if (df := child.minimum_quantity_warning) is not None and not df.empty:
                    stream.write('<p></p>')
                    stream.write('<p>Артикулы, количество которых в спецификации ниже минимального:</p>')
                    stream.write(df.to_html(index=False, justify='center'))

                if (df := child.multiplicity_warning) is not None and not df.empty:
                    stream.write('<p>Артикулы, количество которых не соответствует кратности:</p>')
                    stream.write(df.to_html(index=False, justify='center'))

                if (df := child.warning_999999_price) is not None and not df.empty:
                    stream.write('<p>Артикулы c ценой 999 999. Возможно, стоит подобрать  другие:</p>')
                    stream.write(df.to_html(index=False, justify='center'))

                if (df := child.discount_margin_relation_merged_df) is not None and not df.empty:
                    stream.write('<p>Маржинальность по всей спецификации:</p>')
                    stream.write(df.to_html(index=False, justify='center'))

                if (df := child.price_by_type_of_equipment) is not None and not df.empty:
                    stream.write('<p>Стоимость по типам оборудования:</p>')
                    stream.write(df.to_html(index=False, justify='center'))

                if (data := child.discount_margin_relation_grouped_df) is not None:
                    stream.write('<p>Маржинальность каждого типа оборудования отдельно:</p>')
                    for k, df in data.items():
                        stream.write(f'{k}:')
                        stream.write(df.to_html(index=False, justify='center'))

                if (df := child.equal_margin_df) is not None and not df.empty:
                    files = self._equal_margin_xls.pop(0)
                    for file in files:
                        with pd.ExcelWriter(str(file), engine='openpyxl', mode='w') as writer:
                            df.to_excel(writer, sheet_name='Сводная', index=False, startrow=3)
                        _fill_cells_for_eq_margin(self._current_message.sender.company, file)

                if (df := child.bad_articles) is not None and not df.empty:
                    stream.write('А вот плохие артикулы:')
                    stream.write(df.to_html(index=False, justify='center'))


class ConcatenatedUserResultHandler(IResultHandler):
    _handled_result_cls = ConcatenatedUserResult

    def __init__(self, user_success_handler: SuccessUserResultHandler, bad_handler: BadResultHandler):
        super().__init__()
        self._user_success_handler = user_success_handler
        self._bad_handler = bad_handler

    def _handle(self, result: ConcatenatedUserResult):
        file = self._user_success_handler.reply_file

        with file.open('w', encoding='utf-8') as stream:
            stream.write(f'<header>{self._current_message.receiver.name}</header>')

        if result.accepted_user_result:
            sub_message = self._current_message
            sub_message.data = result.accepted_user_result
            self._user_success_handler.handle(sub_message)
        if result.bad_user_result:
            sub_message = self._current_message
            sub_message.data = result.bad_user_result
            self._bad_handler.handle(sub_message)

        with file.open('a', encoding='utf-8') as stream:
            stream.write(f'<p>C уважением,</p>')
            stream.write(f'<p>Робот расчёта маржи</p>')


def _fill_cells_for_eq_margin(company, file):
    wb = load_workbook(file)
    ws = wb["Сводная"]
    ws['E2'] = "45%"
    ws['D2'] = "Укажите маржу:"
    ws['A1'] = company

    equal_DRM_column_list = [ws['B:B'], ws['D:D'], ws['F:F']]
    for column in equal_DRM_column_list:
        for cell in column:
            cell.number_format = u'#,##0.00€'

    for cell in ws['C:C']:
        cell.number_format = u'##,##%'
    for cell in ws['E:E']:
        cell.number_format = u'##,##%'

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 19
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 14

    for column in equal_DRM_column_list:
        for cell in column:
            cell.alignment = Alignment(horizontal='center')

    for col in ws.iter_cols(min_row=ws.max_row, max_col=6, max_row=ws.max_row):
        for cell in col:
            cell.font = Font(bold=True)
    wb.save(file)


class NoExcelFilesHandler(IResultHandler):
    _handled_result_cls = NoExcelFilesResult

    def __init__(self, output: Path):
        super().__init__()
        self._output = output

    def _handle(self, result: NoExcelFilesResult):
        with self._output.open('w', encoding='utf-8') as stream:
            stream.write(f'<header>{self._current_message.sender.name},</header>')
            stream.write(f'<p>Нет приложений с расширением .xlsx.</p>')
            stream.write(f'<p>C уважением,</p>')
            stream.write(f'<p>Робот расчёта маржи</p>')

